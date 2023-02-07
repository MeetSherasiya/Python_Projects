from tkinter import *
from tkinter.ttk import Combobox
from tkinter import messagebox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib


root = Tk()
root.title("Data Entry")
root.geometry("700x420+300+200")
root.resizable(False,False)
root.configure(bg="#326273")

file = pathlib.Path('Beginners_projects/Data_Entry_Form/Data.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Full Name"
    sheet['B1'] = "Phone Number"
    sheet['C1'] = "Age"
    sheet['D1'] = "Gender"
    sheet['E1'] = "Address"

    file.save('Beginners_projects/Data_Entry_Form/Data.xlsx')

def submit():
    name = name_value.get()
    contact = contact_value.get()
    age = age_value.get()
    gender = gender_entry.get()
    address = address_entry.get(1.0, END)

    file = openpyxl.load_workbook('Beginners_projects/Data_Entry_Form/Data.xlsx')
    sheet = file.active
    sheet.cell(column=1, row= sheet.max_row+1, value=name)
    sheet.cell(column=2, row= sheet.max_row, value=contact)
    sheet.cell(column=3, row= sheet.max_row, value=age)
    sheet.cell(column=4, row= sheet.max_row, value=gender)
    sheet.cell(column=5, row= sheet.max_row, value=address)

    file.save(r'Beginners_projects/Data_Entry_Form/Data.xlsx')


def clear():
    name_value.set('')
    contact_value.set('')
    age_value.set('')
    address_entry.delete(1.0, END)

#******* icon ***********
icon_image = PhotoImage(file="Beginners_projects/Data_Entry_Form/userman.png")
root.iconphoto(False, icon_image)

#******* heading ************
Label(root, text="Please fill out this entry form :", font="arial 23", bg="#326273", fg="#fff").place(x=20, y=20)


#********* Label **********
Label(root, text="Name", font=20, bg="#326273", fg="white").place(x=20, y=80)
Label(root, text="Contact No.", font=20, bg="#326273", fg="white").place(x=20, y=130)
Label(root, text="Age", font=20, bg="#326273", fg="white").place(x=20, y=180)
Label(root, text="Gender", font=20, bg="#326273", fg="white").place(x=20, y=230)
Label(root, text="Address", font=20, bg="#326273", fg="white").place(x=20, y=280)


#*********** Entry ************
name_value = StringVar()
contact_value = StringVar()
age_value = StringVar()

name_entry = Entry(root, textvariable=name_value, width=40, bd=2, font=16)
contact_entry = Entry(root, textvariable=contact_value, width=40, bd=2, font=16)
age_entry = Entry(root, textvariable=age_value, width=40, bd=2, font=16)

gender_entry = Combobox(root, values=['Male','Female','Other'], font="arial 14", state='r', width=14)
gender_entry.set('Male')

address_entry = Text(root, width=50, height=4, bd=4)

name_entry.place(x=150, y=80)
contact_entry.place(x=150, y=130)
age_entry.place(x=150, y=180)
gender_entry.place(x=150, y=230)
address_entry.place(x=150, y=280)

Button(root, text="Submit", bg="#326273", fg="white", width=15, height=2, command=submit).place(x=20, y=370)
Button(root, text="Clear", bg="#326273", fg="white", width=15, height=2, command=clear).place(x=160, y=370)
Button(root, text="Exit", bg="#326273", fg="white", width=15, height=2, command=lambda:root.destroy()).place(x=300, y=370)


root.mainloop()